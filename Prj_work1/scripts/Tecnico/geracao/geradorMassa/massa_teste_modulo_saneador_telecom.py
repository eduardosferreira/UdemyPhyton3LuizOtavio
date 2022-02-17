#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

"""
--------------------------------------------------------------------------------------------------------------------------
  SISTEMA ..: 
  MODULO ...: 
  SCRIPT ...: Criação Massa de Teste
  CRIACAO ..: 13/08/2021
  AUTOR ....: Victor Santos - Kyros Consultoria 
  DESCRICAO : Desenvolver um processo python que gere massa de dados para teste.
  ALTERACAO :
--------------------------------------------------------------------------------------------------------------------------
    Exemplo de comando: ./criacao_massa_teste.py <MESANO> <SERIE> <UF> <FILI_COD> <NUM_NOTA_INI> <NUM_NOTA_FIM> <OWNER> 
--------------------------------------------------------------------------------------------------------------------------
"""

import sys
import datetime
import os
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.reader.excel import load_workbook 

dir_base = os.path.join( os.path.realpath('.').split('/PROD/')[0], 'PROD') if os.path.realpath('.').__contains__('/PROD/') else os.path.join( os.path.realpath('.').split('/DEV/')[0], 'DEV')
sys.path.append(dir_base)
import configuracoes
import comum
import sql

log.gerar_log_em_arquivo = True

config = comum.carregaConfiguracoes(configuracoes) 

SD = ('/' if os.name == 'posix' else '\\')

def processar():

    vDataIni    = ""
    vDataFim    = ""
    vSerie      = ""
    vUF         = ""
    vFiliCod    = ""
    vNumNotaIni = ""   
    vNumNotaFim = "" 
    vOwner      = ""  
    ret         = 0
    
    conexao = sql.geraCnxBD(configuracoes)

    if ( len(sys.argv) == 9 ):
# <MESANO> <SERIE> <UF> <FILI_COD> <NUM_NOTA_INI> <NUM_NOTA_FIM> <OWNER>         
        vDataIni    = sys.argv[1]
        vDataFim    = sys.argv[2]
        vSerie      = sys.argv[3] 
        vUF         = sys.argv[4].upper()
        vFiliCod    = sys.argv[5] 
        vNumNotaIni = sys.argv[6] 
        vNumNotaFim = sys.argv[7] 
        vOwner      = sys.argv[8] 

        if vOwner == '':
            log('#### ERRO - O PARAMETRO OWNER É OBRIGATÓRIO...')
            return(99)

        # if vPeriodo != '':
        #     vDataIni ='01' + str(vPeriodo)
        #     vAno = str(vPeriodo[2:6])
        #     vMes = str(vPeriodo[0:2])
        #     vDataIni ='01/' + vMes +'/'+ vAno
        #     UltDiaMes =ultimodia(int(vPeriodo[2:6]), int(vPeriodo[0:2]))
        #     vDataFim =str(UltDiaMes)+ '/' + vMes + '/' + vAno 


        log("-"* 100)
        log('# - Data Inicial.................................:', vDataIni)
        log('# - Data Fim.....................................:', vDataFim)
        log('# - Série........................................:', vSerie)
        log('# - UF...........................................:', vUF)
        log('# - Fili_cod.....................................:', vFiliCod)
        log('# - Número de Nota Inicial.......................:', vNumNotaIni)
        log('# - Número de Nota Final.........................:', vNumNotaFim)
        log('# - Owner........................................:', vOwner)
        log("-"* 100)
        
    else:
        log("-" * 100)
        log("#### ")
        log('#### ERRO - Erro nos parametros do script.')
        log("Retorno = 99") 
        ret = 99
        return(99)  

#### DADOS
#### DADOS
#### DADOS

    vWhere = '1 = 1'
    
    if vSerie != '':
        result = vSerie.split(',')
        serie = "('"
        for row in result:
            serie = serie + row.replace(' ', '') + "','"
        serie = serie[:len(serie)-2]
        serie = serie + ")"

        vWhere = vWhere + " AND replace(l.serie,' ','') in " + serie

    if vUF != '':
        result = vUF.split(',')
        uf = "('"
        for row in result:
            uf = uf + row.replace(' ', '') + "','"
        uf = uf[:len(uf)-2]
        uf = uf + ")"
        vWhere = vWhere + " AND l.uf_filial in " + uf

    if vDataIni != '':
        vWhere = vWhere + " AND l.mes_ano >= to_date ('" + vDataIni + "', 'dd/mm/yyyy') "
        vWhere = vWhere + " AND l.mes_ano <= to_date ('" + vDataFim + "', 'dd/mm/yyyy') "

    if vFiliCod != '':
        result = vFiliCod.split(',')
        fili = "('"
        for row in result:
            fili = fili + row.replace(' ', '') + "','"
        fili = fili[:len(fili)-2]
        fili = fili + ")"
        vWhere = vWhere + " AND l.fili_cod  in " + fili

    if vNumNotaIni != '':
        vWhere = vWhere + " AND variavel >= '" + vNumNotaIni.zfill(9) + "'"

    if vNumNotaFim != '':
        vWhere = vWhere + " AND variavel <=  '" + vNumNotaFim.zfill(9) + "')"
    
    dados = create_table(vOwner,conexao,vWhere)

    if dados != 0:
        return (dados)
    return(dados)

def validauf(uf):
    return(True if (uf.upper() in ('AC','AL','AM','AP','BA','CE','DF','ES','GO','MA','MG','MS','MT','PA','PB','PE','PI','PR','RJ','RN','RO','RR','RS','SC','SE','SP','TO')) else False)
           
def dtf():
    return (datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
 
def ultimodia(ano,mes):
   return(31 if mes == 12 else datetime.date.fromordinal((datetime.date(ano,mes+1,1)).toordinal()-1).day)

def create_table(vOwner,connection,pWhere):

############ MESTRE_NFTL_SERV_PROD
############ MESTRE_NFTL_SERV_PROD
############ MESTRE_NFTL_SERV_PROD

    tabela = 'MESTRE_NFTL_SERV_PROD'

    vWhere = re.sub( r'variavel', 'm.mnfst_num',  pWhere)

    con=sql.geraCnxBD(configuracoes)

    log(dtf(), '# INICIANDO O PROCESSO DE CRIAÇÃO DA TABELA', tabela)
    log('')
    log(dtf(), '# TENTANDO DROPAR A TABELA', tabela)
    log('')

    query="""
        DROP TABLE %s.%s PURGE
    """%(vOwner, tabela)
    try:
        con.executa(query)        
        log(dtf(), '# TABELA DROPADA COM SUCESSO', tabela)
        log('')
    except Exception as e:
        log(dtf(), '#### TABELA NÃO EXISTE NO OWNER INFORMADO...')
        log('')    

    log(dtf(), '# INICIANDO O CREATE DA TABELA', tabela)
    log('')
    query="""
        CREATE TABLE %s.%s TABLESPACE CADASTRO_DATA as (
            SELECT /*+ PARALLEL (16) */ m.* 
              FROM openrisow.MESTRE_NFTL_SERV_PROD m
              JOIN gfcarga.TSH_SERIE_LEVANTAMENTO l
                ON m.id_serie_levantamento = l.id_serie_levantamento
             WHERE %s
    """%(vOwner,tabela,vWhere)

    log(query)
    try:
        con.executa(query)
        log(dtf(), '# TABELA CRIADA COM SUCESSO', tabela)
        log('')
    except Exception as e:
        log('ERRO ASSINALADO: ', str(e))
        return(99)
    
    log(dtf(), '# INICIANDO CRIAÇÃO DE PK DA TABELA', tabela)
    log('')
    query="""
        ALTER TABLE %s.%s ADD CONSTRAINT PK_%s PRIMARY KEY (MNFST_DTEMISS,MNFST_SERIE,MNFST_NUM,FILI_COD,MDOC_COD,EMPS_COD)
    """%(vOwner,tabela,tabela)
    try:
        con.executa(query)
        log(dtf(), '# PK CRIADA COM SUCESSO PARA A TABELA', tabela)
        log('')
    except Exception as e:
        log('ERRO', str(e))
        return(99)

############ ITEM_NFTL_SERV_PROD
############ ITEM_NFTL_SERV_PROD
############ ITEM_NFTL_SERV_PROD

    tabela = 'ITEM_NFTL_SERV_PROD'

    vWhere = re.sub( r'variavel', 'i.infst_num',  pWhere)

    query="""
        DROP TABLE %s.%s PURGE
    """%(vOwner, tabela)
    try:
        con.executa(query)
        log(dtf(), '# TABELA DROPADA COM SUCESSO', tabela)
        log('')
    except Exception as e:
        log(dtf(), '#### TABELA NÃO EXISTE NO OWNER INFORMADO...')
        log('')    

    log(dtf(), '# INICIANDO O CREATE DA TABELA', tabela)
    log('')
    query="""
        CREATE TABLE %s.%s TABLESPACE CADASTRO_DATA as (
            SELECT /*+ PARALLEL (16) */ i.* 
              FROM openrisow.ITEM_NFTL_SERV_PROD i
              JOIN gfcarga.TSH_SERIE_LEVANTAMENTO l
                ON i.id_serie_levantamento = l.id_serie_levantamento
             WHERE %s
    """%(vOwner,tabela,vWhere)

    log(query)
    try:
        con.executa(query)
        log(dtf(), '# TABELA CRIADA COM SUCESSO', tabela)
        log('')
    except Exception as e:
        log('ERRO ASSINALADO: ', str(e))
        return(99)
    
    log(dtf(), '# INICIANDO CRIAÇÃO DE PK DA TABELA', tabela)
    log('')
    query="""
        ALTER TABLE %s.%s ADD CONSTRAINT PK_%s PRIMARY KEY (INFST_NUM,INFST_DTEMISS,INFST_SERIE,INFST_NUM_SEQ,EMPS_COD,FILI_COD)
    """%(vOwner,tabela,tabela)
    try:
        con.executa(query)
        log(dtf(), '# PK CRIADA COM SUCESSO PARA A TABELA', tabela)
        log('')
    except Exception as e:
        log('ERRO', str(e))
        return(99)

############ BILLING_COMBINADO_FINAL
############ BILLING_COMBINADO_FINAL
############ BILLING_COMBINADO_FINAL

    tabela = 'BILLING_COMBINADO_FINAL'

    vWhere = re.sub( r'variavel', 'b.mnfst_num',  pWhere)

    query="""
        DROP TABLE %s.%s PURGE
    """%(vOwner, tabela)
    try:
        con.executa(query)
        log(dtf(), '# TABELA DROPADA COM SUCESSO', tabela)
        log('')
    except Exception as e:
        log(dtf(), '#### TABELA NÃO EXISTE NO OWNER INFORMADO...')
        log('')    

    log(dtf(), '# INICIANDO O CREATE DA TABELA', tabela)
    log('')
    query="""
        CREATE TABLE %s.%s TABLESPACE CADASTRO_DATA as (
            SELECT /*+ PARALLEL (16) */ b.* 
              FROM gfcadastro.BILLING_COMBINADO_FINAL b
              JOIN gfcarga.TSH_SERIE_LEVANTAMENTO l
                ON l.emps_cod              = b.emps_cod
               AND l.uf_filial             = b.uf_filial
               AND replace(l.serie,' ','') = replace(b.mnfst_serie,' ','')
               AND l.mes_ano               = to_date('01' || to_char(b.mnfst_dtemiss,'MMYYYY'),'DDMMYYYY')
             WHERE %s
    """%(vOwner,tabela,vWhere)

    log(query)
    try:
        con.executa(query)
        log(dtf(), '# TABELA CRIADA COM SUCESSO', tabela)
        log('')
    except Exception as e:
        log('ERRO ASSINALADO: ', str(e))
        return(99)
    
    log(dtf(), '# INICIANDO CRIAÇÃO DE PK DA TABELA', tabela)
    log('')
    query="""
        ALTER TABLE %s.%s ADD CONSTRAINT PK_%s PRIMARY KEY (MNFST_NUM,MNFST_SERIE,MNFST_DTEMISS,FILI_COD,MDOC_COD,EMPS_COD,ORIGEM,UF_FILIAL)
    """%(vOwner,tabela,tabela)
    try:
        con.executa(query)
        log(dtf(), '# PK CRIADA COM SUCESSO PARA A TABELA', tabela)
        log('')
    except Exception as e:
        log('ERRO', str(e))
        return(99)

############ TSH_MESTRE_CONV_115
############ TSH_MESTRE_CONV_115
############ TSH_MESTRE_CONV_115

    tabela = 'TSH_MESTRE_CONV_115'

    vWhere = re.sub( r'variavel', 'm.numero_NF',  pWhere)

    query="""
        DROP TABLE %s.%s PURGE
    """%(vOwner, tabela)
    try:
        con.executa(query)
        log(dtf(), '# TABELA DROPADA COM SUCESSO', tabela)
        log('')
    except Exception as e:
        log(dtf(), '#### TABELA NÃO EXISTE NO OWNER INFORMADO...')
        log('')    

    log(dtf(), '# INICIANDO O CREATE DA TABELA', tabela)
    log('')
    query="""
        CREATE TABLE %s.%s TABLESPACE CADASTRO_DATA as (
            SELECT  /*+ PARALLEL (16) */ m.*   
              FROM gfcarga.TSH_MESTRE_CONV_115 m
              JOIN gfcarga.TSH_SERIE_LEVANTAMENTO    l
                ON l.id_serie_levantamento = m.id_serie_levantamento
               AND l.uf_filial             = m.uf_filial
             WHERE %s
    """%(vOwner,tabela,vWhere)

    log(query)
    try:
        con.executa(query)
        log(dtf(), '# TABELA CRIADA COM SUCESSO', tabela)
        log('')
    except Exception as e:
        log('ERRO ASSINALADO: ', str(e))
        return(99)
    
    log(dtf(), '# INICIANDO CRIAÇÃO DE PK DA TABELA', tabela)
    log('')
    query="""
        ALTER TABLE %s.%s ADD CONSTRAINT PK_%s PRIMARY KEY (ID_SERIE_LEVANTAMENTO,UF_FILIAL,NUMERO_NF)
    """%(vOwner,tabela,tabela)
    try:
        con.executa(query)
        log(dtf(), '# PK CRIADA COM SUCESSO PARA A TABELA', tabela)
        log('')
    except Exception as e:
        log('ERRO', str(e))
        return(99)

############ TSH_DESTINATARIO_CONV_115
############ TSH_DESTINATARIO_CONV_115
############ TSH_DESTINATARIO_CONV_115

    tabela = 'TSH_DESTINATARIO_CONV_115'

    vWhere = re.sub( r'variavel', 'm.numero_NF',  pWhere)

    query="""
        DROP TABLE %s.%s PURGE
    """%(vOwner, tabela)
    try:
        con.executa(query)
        log(dtf(), '# TABELA DROPADA COM SUCESSO', tabela)
        log('')
    except Exception as e:
        log(dtf(), '#### TABELA NÃO EXISTE NO OWNER INFORMADO...')
        log('')    

    log(dtf(), '# INICIANDO O CREATE DA TABELA', tabela)
    log('')
    query="""
        CREATE TABLE %s.%s TABLESPACE CADASTRO_DATA as (
            SELECT /*+ PARALLEL (16) */ d.*   
              FROM gfcarga.TSH_MESTRE_CONV_115 m
              JOIN gfcarga.TSH_SERIE_LEVANTAMENTO    l
                ON l.id_serie_levantamento = m.id_serie_levantamento
               AND l.uf_filial             = m.uf_filial
              JOIN gfcarga.TSH_DESTINATARIO_CONV_115 d
                ON m.ID_SERIE_LEVANTAMENTO = d.ID_SERIE_LEVANTAMENTO
               AND m.UF_FILIAL             = d.UF_FILIAL
               AND m.LINHA                 = d.LINHA
               AND m.VOLUME                = d.VOLUME
             WHERE %s
    """%(vOwner,tabela,vWhere)

    log(query)
    try:
        con.executa(query)
        log(dtf(), '# TABELA CRIADA COM SUCESSO', tabela)
        log('')
    except Exception as e:
        log('ERRO ASSINALADO: ', str(e))
        return(99)
    
    log(dtf(), '# INICIANDO CRIAÇÃO DE PK DA TABELA', tabela)
    log('')
    query="""
        ALTER TABLE %s.%s ADD CONSTRAINT PK_%s PRIMARY KEY (ID_SERIE_LEVANTAMENTO,UF_FILIAL,LINHA)
    """%(vOwner,tabela,tabela)
    try:
        con.executa(query)
        log(dtf(), '# PK CRIADA COM SUCESSO PARA A TABELA', tabela)
        log('')
    except Exception as e:
        log('ERRO', str(e))
        return(99)

############ TSH_ITEM_CONV_115
############ TSH_ITEM_CONV_115
############ TSH_ITEM_CONV_115

    tabela = 'TSH_ITEM_CONV_115'

    vWhere = re.sub( r'variavel', 'm.numero_NF',  pWhere)

    query="""
        DROP TABLE %s.%s PURGE
    """%(vOwner, tabela)
    try:
        con.executa(query)
        log(dtf(), '# TABELA DROPADA COM SUCESSO', tabela)
        log('')
    except Exception as e:
        log(dtf(), '#### TABELA NÃO EXISTE NO OWNER INFORMADO...')
        log('')    

    log(dtf(), '# INICIANDO O CREATE DA TABELA', tabela)
    log('')
    query="""
        CREATE TABLE %s.%s TABLESPACE CADASTRO_DATA as (
            SELECT /*+ PARALLEL (16) */ i.*   
              FROM gfcarga.TSH_MESTRE_CONV_115 m
              JOIN gfcarga.TSH_SERIE_LEVANTAMENTO    l
                ON l.id_serie_levantamento = m.id_serie_levantamento
               AND l.uf_filial             = m.uf_filial
              JOIN gfcarga.TSH_ITEM_CONV_115         i
                ON m.ID_SERIE_LEVANTAMENTO = i.ID_SERIE_LEVANTAMENTO
               AND m.UF_FILIAL             = i.UF_FILIAL
               AND m.NUMERO_NF             = i.NUMERO_NF
             WHERE %s
    """%(vOwner,tabela,vWhere)

    log(query)
    try:
        con.executa(query)
        log(dtf(), '# TABELA CRIADA COM SUCESSO', tabela)
        log('')
    except Exception as e:
        log('ERRO ASSINALADO: ', str(e))
        return(99)
    
    log(dtf(), '# INICIANDO CRIAÇÃO DE PK DA TABELA', tabela)
    log('')
    query="""
        ALTER TABLE %s.%s ADD CONSTRAINT PK_%s PRIMARY KEY (ID_SERIE_LEVANTAMENTO,UF_FILIAL,NUMERO_NF,NUM_ITEM)
    """%(vOwner,tabela,tabela)
    try:
        con.executa(query)
        log(dtf(), '# PK CRIADA COM SUCESSO PARA A TABELA', tabela)
        log('')
    except Exception as e:
        log('ERRO', str(e))
        return(99)
    return(0)

if __name__ == "__main__":
    log("-"*70)
    log("#### ",dtf(), " INICIO DO PROCESSO DE CRIAÇÃO DE MASSA PARA TESTES... ####")
    ret = processar()
    if (ret > 0) :
        log("#### Código de execução = ", ret)
    log("")
    log("#### ",dtf(), " FIM DO PROCESSO... ####")
    sys.exit(ret) 
